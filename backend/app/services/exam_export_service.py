"""Exam results export service."""
import io
import logging

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from app.models.exam import Exam, ExamSession

logger = logging.getLogger(__name__)


class ExamExportService:
    """Service for exporting exam results to PDF and Excel."""

    def generate_pdf_report(
        self,
        exam: Exam,
        sessions: list[ExamSession],
        include_answers: bool = False,
    ) -> bytes:
        """
        Generate PDF report for exam results.

        Args:
            exam: The exam
            sessions: List of exam sessions
            include_answers: Whether to include individual answers

        Returns:
            PDF file bytes
        """
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        story = []
        styles = getSampleStyleSheet()

        # Custom styles
        title_style = ParagraphStyle(
            "CustomTitle",
            parent=styles["Heading1"],
            fontSize=24,
            textColor=colors.HexColor("#4F46E5"),
            spaceAfter=30,
        )

        heading_style = ParagraphStyle(
            "CustomHeading",
            parent=styles["Heading2"],
            fontSize=16,
            textColor=colors.HexColor("#1F2937"),
            spaceAfter=12,
        )

        # Title
        story.append(Paragraph(f"Exam Results Report: {exam.title}", title_style))
        story.append(Spacer(1, 0.2 * inch))

        # Exam information
        story.append(Paragraph("Exam Information", heading_style))
        exam_info = [
            ["Exam ID:", str(exam.id)],
            ["Type:", exam.exam_type],
            ["Status:", exam.status],
            [
                "Time Limit:",
                f"{exam.time_limit_minutes} minutes"
                if exam.time_limit_minutes
                else "No limit",
            ],
            ["Max Attempts:", str(exam.max_attempts)],
            [
                "Passing Score:",
                f"{exam.passing_score}%" if exam.passing_score else "Not set",
            ],
            [
                "Created:",
                exam.created_at.strftime("%Y-%m-%d %H:%M")
                if exam.created_at
                else "N/A",
            ],
        ]

        info_table = Table(exam_info, colWidths=[2 * inch, 4 * inch])
        info_table.setStyle(
            TableStyle(
                [
                    ("FONT", (0, 0), (-1, -1), "Helvetica", 10),
                    ("FONT", (0, 0), (0, -1), "Helvetica-Bold", 10),
                    ("TEXTCOLOR", (0, 0), (-1, -1), colors.black),
                    ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#F3F4F6")),
                ]
            )
        )
        story.append(info_table)
        story.append(Spacer(1, 0.3 * inch))

        # Summary statistics
        completed_sessions = [s for s in sessions if s.status == "completed"]
        total_sessions = len(sessions)
        total_completed = len(completed_sessions)

        if total_completed > 0:
            avg_score = (
                sum(s.score for s in completed_sessions if s.score) / total_completed
            )
            passed_sessions = [
                s
                for s in completed_sessions
                if s.score and exam.passing_score and s.score >= exam.passing_score
            ]
            pass_rate = (
                (len(passed_sessions) / total_completed * 100)
                if exam.passing_score
                else 0
            )
        else:
            avg_score = 0
            pass_rate = 0

        story.append(Paragraph("Summary Statistics", heading_style))
        summary_data = [
            ["Total Sessions:", str(total_sessions)],
            ["Completed Sessions:", str(total_completed)],
            [
                "Completion Rate:",
                f"{(total_completed / total_sessions * 100):.1f}%"
                if total_sessions > 0
                else "0%",
            ],
            ["Average Score:", f"{avg_score:.2f}%" if total_completed > 0 else "N/A"],
            ["Pass Rate:", f"{pass_rate:.2f}%" if exam.passing_score else "N/A"],
        ]

        summary_table = Table(summary_data, colWidths=[2 * inch, 4 * inch])
        summary_table.setStyle(
            TableStyle(
                [
                    ("FONT", (0, 0), (-1, -1), "Helvetica", 10),
                    ("FONT", (0, 0), (0, -1), "Helvetica-Bold", 10),
                    ("TEXTCOLOR", (0, 0), (-1, -1), colors.black),
                    ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#F3F4F6")),
                ]
            )
        )
        story.append(summary_table)
        story.append(Spacer(1, 0.3 * inch))

        # Session results
        if completed_sessions:
            story.append(Paragraph("Detailed Results", heading_style))

            # Table headers
            session_data = [
                [
                    "Session ID",
                    "Candidate ID",
                    "Status",
                    "Score",
                    "Time Taken",
                    "Completed At",
                ]
            ]

            # Add session rows
            for session in completed_sessions:
                time_taken = "N/A"
                if session.started_at and session.completed_at:
                    duration = session.completed_at - session.started_at
                    time_taken = f"{duration.total_seconds() / 60:.1f} min"

                session_data.append(
                    [
                        str(session.id),
                        str(session.candidate_id),
                        session.status,
                        f"{session.score:.1f}%" if session.score else "N/A",
                        time_taken,
                        session.completed_at.strftime("%Y-%m-%d %H:%M")
                        if session.completed_at
                        else "N/A",
                    ]
                )

            results_table = Table(
                session_data,
                colWidths=[
                    0.8 * inch,
                    1 * inch,
                    0.9 * inch,
                    0.8 * inch,
                    1 * inch,
                    1.5 * inch,
                ],
            )
            results_table.setStyle(
                TableStyle(
                    [
                        ("FONT", (0, 0), (-1, 0), "Helvetica-Bold", 9),
                        ("FONT", (0, 1), (-1, -1), "Helvetica", 8),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4F46E5")),
                        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                        (
                            "ROWBACKGROUNDS",
                            (0, 1),
                            (-1, -1),
                            [colors.white, colors.HexColor("#F9FAFB")],
                        ),
                    ]
                )
            )
            story.append(results_table)

        # Build PDF
        doc.build(story)
        pdf_bytes = buffer.getvalue()
        buffer.close()

        return pdf_bytes

    def generate_excel_report(
        self,
        exam: Exam,
        sessions: list[ExamSession],
        include_answers: bool = False,
    ) -> bytes:
        """
        Generate Excel report for exam results.

        Args:
            exam: The exam
            sessions: List of exam sessions
            include_answers: Whether to include individual answers

        Returns:
            Excel file bytes
        """
        workbook = Workbook()

        # Remove default sheet
        if "Sheet" in workbook.sheetnames:
            workbook.remove(workbook["Sheet"])

        # Summary sheet
        summary_sheet = workbook.create_sheet("Summary")
        self._create_summary_sheet(summary_sheet, exam, sessions)

        # Results sheet
        results_sheet = workbook.create_sheet("Detailed Results")
        self._create_results_sheet(results_sheet, exam, sessions)

        # Save to bytes
        buffer = io.BytesIO()
        workbook.save(buffer)
        excel_bytes = buffer.getvalue()
        buffer.close()

        return excel_bytes

    def _create_summary_sheet(self, sheet, exam: Exam, sessions: list[ExamSession]):
        """Create summary sheet in Excel workbook."""
        # Header styling
        PatternFill(
            start_color="4F46E5", end_color="4F46E5", fill_type="solid"
        )
        Font(color="FFFFFF", bold=True, size=14)

        # Title
        sheet["A1"] = "Exam Results Summary"
        sheet["A1"].font = Font(bold=True, size=16, color="4F46E5")
        sheet.merge_cells("A1:B1")

        # Exam information
        sheet["A3"] = "Exam Information"
        sheet["A3"].font = Font(bold=True, size=12)

        exam_info = [
            ["Exam ID", exam.id],
            ["Title", exam.title],
            ["Type", exam.exam_type],
            ["Status", exam.status],
            [
                "Time Limit",
                f"{exam.time_limit_minutes} minutes"
                if exam.time_limit_minutes
                else "No limit",
            ],
            ["Max Attempts", exam.max_attempts],
            [
                "Passing Score",
                f"{exam.passing_score}%" if exam.passing_score else "Not set",
            ],
            [
                "Created",
                exam.created_at.strftime("%Y-%m-%d %H:%M")
                if exam.created_at
                else "N/A",
            ],
        ]

        row = 4
        for label, value in exam_info:
            sheet[f"A{row}"] = label
            sheet[f"B{row}"] = value
            sheet[f"A{row}"].font = Font(bold=True)
            row += 1

        # Statistics
        completed_sessions = [s for s in sessions if s.status == "completed"]
        total_sessions = len(sessions)
        total_completed = len(completed_sessions)

        if total_completed > 0:
            avg_score = (
                sum(s.score for s in completed_sessions if s.score) / total_completed
            )
            passed_sessions = [
                s
                for s in completed_sessions
                if s.score and exam.passing_score and s.score >= exam.passing_score
            ]
            pass_rate = (
                (len(passed_sessions) / total_completed * 100)
                if exam.passing_score
                else 0
            )
        else:
            avg_score = 0
            pass_rate = 0

        sheet[f"A{row + 1}"] = "Statistics"
        sheet[f"A{row + 1}"].font = Font(bold=True, size=12)

        stats = [
            ["Total Sessions", total_sessions],
            ["Completed Sessions", total_completed],
            [
                "Completion Rate",
                f"{(total_completed / total_sessions * 100):.1f}%"
                if total_sessions > 0
                else "0%",
            ],
            ["Average Score", f"{avg_score:.2f}%" if total_completed > 0 else "N/A"],
            ["Pass Rate", f"{pass_rate:.2f}%" if exam.passing_score else "N/A"],
        ]

        row += 2
        for label, value in stats:
            sheet[f"A{row}"] = label
            sheet[f"B{row}"] = value
            sheet[f"A{row}"].font = Font(bold=True)
            row += 1

        # Adjust column widths
        sheet.column_dimensions["A"].width = 20
        sheet.column_dimensions["B"].width = 30

    def _create_results_sheet(self, sheet, exam: Exam, sessions: list[ExamSession]):
        """Create detailed results sheet in Excel workbook."""
        # Headers
        headers = [
            "Session ID",
            "Candidate ID",
            "Status",
            "Score (%)",
            "Time Taken (minutes)",
            "Started At",
            "Completed At",
            "Passed",
        ]

        header_fill = PatternFill(
            start_color="4F46E5", end_color="4F46E5", fill_type="solid"
        )
        header_font = Font(color="FFFFFF", bold=True)

        for col, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Data rows
        completed_sessions = [s for s in sessions if s.status == "completed"]
        for row, session in enumerate(completed_sessions, start=2):
            # Calculate time taken
            time_taken = ""
            if session.started_at and session.completed_at:
                duration = session.completed_at - session.started_at
                time_taken = f"{duration.total_seconds() / 60:.1f}"

            # Check if passed
            passed = "N/A"
            if exam.passing_score and session.score:
                passed = "Yes" if session.score >= exam.passing_score else "No"

            sheet.cell(row=row, column=1, value=session.id)
            sheet.cell(row=row, column=2, value=session.candidate_id)
            sheet.cell(row=row, column=3, value=session.status)
            sheet.cell(
                row=row,
                column=4,
                value=f"{session.score:.1f}" if session.score else "",
            )
            sheet.cell(row=row, column=5, value=time_taken)
            sheet.cell(
                row=row,
                column=6,
                value=session.started_at.strftime("%Y-%m-%d %H:%M")
                if session.started_at
                else "",
            )
            sheet.cell(
                row=row,
                column=7,
                value=session.completed_at.strftime("%Y-%m-%d %H:%M")
                if session.completed_at
                else "",
            )
            sheet.cell(row=row, column=8, value=passed)

        # Adjust column widths
        for col in range(1, len(headers) + 1):
            sheet.column_dimensions[get_column_letter(col)].width = 18


# Singleton instance
exam_export_service = ExamExportService()
